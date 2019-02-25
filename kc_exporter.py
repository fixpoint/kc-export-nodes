# coding: UTF-8

import yaml
import argparse
import os
import csv
import openpyxl
import json
import jmespath
import datetime
import importlib

from collections import defaultdict
from lib.helper import logger
from lib.kompira_cloud import KompiraCloudAPI
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

accept_format = ['csv', 'xlsx']

border = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000')
)

def get_columns(datatype):
    if datatype == 'managed-nodes':
        rule_module = importlib.import_module('lib.column_managed_nodes')
        return rule_module.columns
    else:
        rule_module = importlib.import_module('lib.column_snapshot_nodes')
        return rule_module.columns

def export_csv(rows, filename):
    logger.info("Export node list: csv")
    with open(filename, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def export_xlsx(rows, filename):
    logger.info("Export node list: xlsx")
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('NodeList')

    max_dims = defaultdict(lambda: 0)
    # Headers
    headers = rows[0].keys()
    for i, header in enumerate(headers):
        cell = sheet.cell(row=1, column=i+1, value=header)
        max_dims[header] = len(header)

        cell.border = border

    # Values
    row_num = 2
    for row in rows:
        for i, header in enumerate(headers):
            cell = sheet.cell(row=row_num, column=i+1, value=row[header])

            cell.border = border

            if type(row[header]) == str:
                col_len = len(row[header])
            else:
                col_len = len(str(row[header]))
            if col_len > max_dims[header]:
                max_dims[header] = col_len
        row_num += 1

    # cell size
    for i, header in enumerate(headers):
        if max_dims[header]:
            sheet.column_dimensions[get_column_letter(i+1)].width = max_dims[header]

    if wb['Sheet']:
        wb.remove(wb['Sheet'])
    wb.save(filename)


def main(args):
    logger.info('Args value check')
    if not os.path.exists(args.config_path):
        logger.error('file "%s" is not found.' % args.config_path)
        exit(1)
    target = ''
    if args.url[-14:] == '/managed-nodes':
        target = 'managed-nodes'
    elif args.url[-6:] == '/nodes':
        target = 'snapshot-nodes'
    else:
        logger.error('%s invalid url.' % args.url)
        exit(1)

    if args.format not in accept_format:
        logger.error('format "%s" is not supported.' % args.format)
        exit(1)

    if args.filename == "":
        logger.error('please input filename.' % args.format)
        exit(1)

    with open(args.config_path) as f:
        config = yaml.load(f)

    logger.info('Read config YML')
    if config['kompira_cloud'].get('basic_auth'):
        kcapi = KompiraCloudAPI(
            config['kompira_cloud']['token'],
            username=config['kompira_cloud']['basic_auth']['username'],
            password=config['kompira_cloud']['basic_auth']['password'],
        )
    else:
        kcapi = KompiraCloudAPI(config['kompira_cloud']['token'])

    logger.info("Get list from KompiraCloud")

    res_json = kcapi.get_from_url(args.url, {'limit': 500})
    items = res_json['items']
    logger.debug(items)

    columns = get_columns(target)

    rows = []
    for item in items:
        row = {}
        for key, val in columns.items():
            if 'zeroth' in val:
                v = jmespath.search(val['path_zeroth'], item)
            else:
                v = jmespath.search(val['path'], item)
            if isinstance(v, list) or isinstance(v, dict):
                v = json.dumps(v)
            row[key] = v
        rows.append(row)

    if args.format == 'csv':
        export_csv(rows, '%s.csv' % args.filename)
    elif args.format == 'xlsx':
        export_xlsx(rows, '%s.xlsx' % args.filename)

    logger.info('Output complete.')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--url', type=str, required=True)
    parser.add_argument('--config_path', type=str, default='config.yml')
    parser.add_argument('--filename', type=str)
    parser.add_argument('--format', type=str, default='csv')
    parser.add_argument('--zeroth', action='store_true')
    args = parser.parse_args()
    main(args)
