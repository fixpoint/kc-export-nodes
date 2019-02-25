import collections
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

def get_exporter(format):
    if format == 'csv':
        return CSVExporter()
    elif format == 'xlsx':
        return ExcelExporter()
    else:
        return None

class Exporter:
    def __init__(self):
        pass

    def export(self, rows, filename):
        pass

class CSVExporter(Exporter):
    def __init__(self):
        self.format = 'csv'

    def export(self, rows, filename):
        filename += '.csv'
        with open(filename, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

class ExcelExporter(Exporter):
    border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )

    def __init__(self):
        self.format = 'xlsx'

    def export(self, rows, filename):
        filename += '.xlsx'
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet('NodeList')

        max_dims = collections.defaultdict(lambda: 0)
        # Headers
        headers = rows[0].keys()
        for i, header in enumerate(headers):
            cell = sheet.cell(row=1, column=i+1, value=header)
            max_dims[header] = len(header)

            cell.border = ExcelExporter.border

        # Values
        row_num = 2
        for row in rows:
            for i, header in enumerate(headers):
                cell = sheet.cell(row=row_num, column=i+1, value=row[header])

                cell.border = ExcelExporter.border

                if type(row[header]) == str:
                    col_len = len(row[header])
                else:
                    col_len = len(str(row[header]))
                if col_len > max_dims[header]:
                    max_dims[header] = col_len
            row_num += 1

        # cell size
        for i, header in enumerate(headers):
            if max_dims[header]:
                sheet.column_dimensions[get_column_letter(i+1)].width = max_dims[header]

        if wb['Sheet']:
            wb.remove(wb['Sheet'])
        wb.save(filename)
